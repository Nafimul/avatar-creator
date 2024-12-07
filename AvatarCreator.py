#! python3
# AvatarCreator.py - Asks the user to choose between different options and
# creates a png of the corresponding avatar. Then, draws the avatar in excel.

import subprocess, openpyxl, platform
from PIL import Image
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill

BODY_PART_HEIGHT = 25
DEFAULT_CELL_WIDTH = 50.4


def addBodyPart(indexFromTop, partName, option1, option2, option3):
	print(
		"Choose your %s. Enter 1 for %s, 2 for %s, 3 for %s"
		% (partName, option1, option2, option3)
	)
	choice = input()

	if choice == "1" or choice == "2" or choice == "3":
		bodyPartIm = Image.open(Path("assets", partName + choice + ".png"))
		avatarIm.paste(bodyPartIm, (0, indexFromTop * BODY_PART_HEIGHT))
	else:
		print("Not a valid choice")
		addBodyPart(indexFromTop, partName, option1, option2, option3)


def rgbToHex(rgbColor):
	color = "{r:02x}{g:02x}{b:02x}"
	return color.format(r=rgbColor[0], g=rgbColor[1], b=rgbColor[2])


avatarIm = Image.new("RGBA", (50, 100))

addBodyPart(0, "hairstyle", "spiky", "bald", "normal")
addBodyPart(1, "head", "long neck and mustached", "bearded", "pinoccio")
addBodyPart(2, "body", "blocky", "stick", "none")
addBodyPart(3, "legs", "rockets", "two legs", "three legs")

# draw the avatar in excel
workbook = Workbook()
sheet = workbook.active
sheet["A1"] = "ZOOM OUT A LOT"
# loop through every pixel in the image and color the appropriate cell in excel
for y in range(avatarIm.height):
	for x in range(avatarIm.width):
		color = avatarIm.getpixel((x, y))
		sheet[openpyxl.utils.get_column_letter(x + 1) + str(y + 1)].fill = PatternFill(
			"solid", fgColor=rgbToHex(color)
		)

	# make all the cells into squares
	sheet.row_dimensions[
		y + 1
	].height = DEFAULT_CELL_WIDTH

avatarIm.save("avatar.png")
avatarIm.close()
workbook.save("avatar.xlsx")
workbook.close()

if platform.system() == "Windows":
	subprocess.run(["start", "avatar.png"], shell=True)
	subprocess.run(["start", "avatar.xlsx"], shell=True)
elif platform.system() == "Darwin":  # mac
	subprocess.run(["open", "avatar.png"])
	subprocess.run(["open", "avatar.xlsx"])
elif platform.system() == "Linux":
	subprocess.run(["xdg-open", "avatar.png"])
	subprocess.run(["xdg-open", "avatar.xlsx"])
else:
	print(
		"Sorry. We can't open the png and excel spreadsheet automaticaly on your os.\
		You'll have to open them yourself"
	)
